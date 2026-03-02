#!/usr/bin/env python3
"""
Excel (.xlsx) → MediaWiki wikitext 转换器

将 Excel 文件的每个 Sheet 转换为 wikitable 格式，输出为 .wikitext 文件。
支持合并单元格、多行文本、空行跳过。

用法:
    python xlsx_to_wikitext.py input.xlsx
    python xlsx_to_wikitext.py input.xlsx -o output.wikitext
    python xlsx_to_wikitext.py input.xlsx --sheet "卷轴架"  # 只转换指定 sheet
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误: 请先安装 openpyxl")
    print("  pip install openpyxl")
    sys.exit(1)


def cell_to_wiki(value):
    """将单元格值转为 wikitext，保留换行"""
    if value is None:
        return ""
    text = str(value).strip()
    # 换行符转为 <br/>
    text = text.replace("\n", "<br/>")
    return text


def get_merged_cell_value(ws, row, col):
    """获取合并单元格的值（从左上角单元格取值）"""
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
                merged_range.min_col <= col <= merged_range.max_col):
            # 只在左上角返回值，其他位置返回 None 表示被合并
            if row == merged_range.min_row and col == merged_range.min_col:
                val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                rowspan = merged_range.max_row - merged_range.min_row + 1
                colspan = merged_range.max_col - merged_range.min_col + 1
                return val, rowspan, colspan
            else:
                return None, 0, 0  # 被合并的单元格，跳过
    # 不在合并范围内
    val = ws.cell(row=row, column=col).value
    return val, 1, 1


def sheet_to_wikitable(ws, sheet_name):
    """将一个 Sheet 转为 wikitable 字符串"""
    lines = []
    lines.append(f"=== {sheet_name} ===")
    lines.append("")

    # 找到实际使用的列范围（跳过全空列）
    max_col = ws.max_column
    max_row = ws.max_row

    # 如果 sheet 很小或空，跳过
    if max_row < 1:
        return ""

    has_merged = len(ws.merged_cells.ranges) > 0

    if has_merged:
        lines.append('{| class="wikitable" style="width:100%;"')
        for row_idx in range(1, max_row + 1):
            # 检查整行是否为空
            row_vals = []
            skip_cols = set()
            for col_idx in range(1, max_col + 1):
                val, rowspan, colspan = get_merged_cell_value(ws, row_idx, col_idx)
                if rowspan == 0:
                    skip_cols.add(col_idx)
                    continue
                row_vals.append((col_idx, val, rowspan, colspan))

            # 全空行跳过
            if all(v is None or str(v).strip() == "" for _, v, _, _ in row_vals):
                continue

            lines.append("|-")
            for col_idx, val, rowspan, colspan in row_vals:
                text = cell_to_wiki(val)
                attrs = []
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

                attr_str = " ".join(attrs)
                if attr_str:
                    lines.append(f"| {attr_str} | {text}")
                else:
                    lines.append(f"| {text}")
        lines.append("|}")
    else:
        # 无合并单元格，简单处理
        lines.append('{| class="wikitable" style="width:100%;"')
        for row_idx in range(1, max_row + 1):
            cells = []
            for col_idx in range(1, max_col + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                cells.append(cell_to_wiki(val))

            # 全空行跳过
            if all(c == "" for c in cells):
                continue

            # 去掉尾部空单元格
            while cells and cells[-1] == "":
                cells.pop()

            if not cells:
                continue

            lines.append("|-")
            for c in cells:
                lines.append(f"| {c}")

        lines.append("|}")

    lines.append("")
    return "\n".join(lines)


def convert_xlsx(input_path, output_path=None, sheet_filter=None):
    """主转换函数"""
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.with_suffix(".wikitext")
    else:
        output_path = Path(output_path)

    print(f"读取: {input_path}")
    wb = openpyxl.load_workbook(str(input_path), data_only=True)

    parts = []
    page_name = input_path.stem.strip()
    # 去掉临时文件前缀
    if page_name.startswith("~$"):
        page_name = page_name[2:]

    parts.append(f"== {page_name} ==")
    parts.append("")

    for name in wb.sheetnames:
        if sheet_filter and name != sheet_filter:
            continue
        ws = wb[name]
        print(f"  转换 Sheet: {name} ({ws.max_row} 行 x {ws.max_column} 列)")
        result = sheet_to_wikitable(ws, name)
        if result:
            parts.append(result)

    content = "\n".join(parts)
    output_path.write_text(content, encoding="utf-8")
    print(f"输出: {output_path} ({len(content)} 字符)")


def main():
    parser = argparse.ArgumentParser(description="Excel → MediaWiki wikitext 转换器")
    parser.add_argument("input", help="输入的 .xlsx 文件路径")
    parser.add_argument("-o", "--output", help="输出的 .wikitext 文件路径（默认同名）")
    parser.add_argument("--sheet", help="只转换指定的 Sheet")
    args = parser.parse_args()

    convert_xlsx(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()
