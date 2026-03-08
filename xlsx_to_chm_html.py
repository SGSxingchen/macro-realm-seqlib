#!/usr/bin/env python3
"""
Excel (.xlsx) -> HTML 转换器（适配 CHM 阅读）

特性：
- 保留合并单元格（rowspan / colspan）
- 保留字体样式（粗体、斜体、下划线、删除线、颜色、字号、字体族）
- 保留单元格背景色、对齐方式、边框样式
- 保留原始行高列宽
- 支持嵌入图片（data URI）
- 支持作为独立脚本调试，也可被 build_chm.py 导入调用
"""

from __future__ import annotations

import argparse
import base64
import colorsys
import html
from pathlib import Path
from typing import Dict, List, Optional, Tuple


# ============================================================
# Excel 默认主题色（Office 2007-2019 标准主题）
# 索引 0-9，可被工作簿自定义覆盖
# ============================================================
DEFAULT_THEME_COLORS = [
    "FFFFFF",  # 0  lt1  (Window Background)
    "000000",  # 1  dk1  (Window Text)
    "E7E6E6",  # 2  lt2
    "44546A",  # 3  dk2
    "4472C4",  # 4  accent1
    "ED7D31",  # 5  accent2
    "A5A5A5",  # 6  accent3
    "FFC000",  # 7  accent4
    "5B9BD5",  # 8  accent5
    "70AD47",  # 9  accent6
]

# Excel indexed 颜色表（前 64 色）
INDEXED_COLORS = [
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "800000", "008000", "000080", "808000", "800080", "008080", "C0C0C0", "808080",
    "9999FF", "993366", "FFFFCC", "CCFFFF", "660066", "FF8080", "0066CC", "CCCCFF",
    "000080", "FF00FF", "FFFF00", "00FFFF", "800080", "800000", "008080", "0000FF",
    "00CCFF", "CCFFFF", "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99",
    "3366FF", "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",
    "003366", "339966", "003300", "333300", "993300", "993366", "333399", "333333",
]


def _apply_tint(hex_rgb: str, tint: float) -> str:
    """对 RGB 颜色应用 Excel tint（-1.0 到 1.0）"""
    if abs(tint) < 0.001:
        return hex_rgb
    r = int(hex_rgb[0:2], 16) / 255.0
    g = int(hex_rgb[2:4], 16) / 255.0
    b = int(hex_rgb[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    l = max(0.0, min(1.0, l))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l, s)
    return f"{int(r2*255):02X}{int(g2*255):02X}{int(b2*255):02X}"


def _extract_theme_colors(wb) -> List[str]:
    """尝试从工作簿 theme XML 中提取主题色，失败则用默认"""
    try:
        import xml.etree.ElementTree as ET
        theme_xml = wb.loaded_theme
        if not theme_xml:
            return list(DEFAULT_THEME_COLORS)
        ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        root = ET.fromstring(theme_xml)
        theme_el = root.find(".//a:themeElements/a:clrScheme", ns)
        if theme_el is None:
            return list(DEFAULT_THEME_COLORS)
        # Excel 主题色映射顺序: dk1, lt1, dk2, lt2, accent1~6
        tag_order = ["dk1", "lt1", "dk2", "lt2",
                     "accent1", "accent2", "accent3", "accent4", "accent5", "accent6"]
        colors = list(DEFAULT_THEME_COLORS)
        for i, tag in enumerate(tag_order):
            el = theme_el.find(f"a:{tag}", ns)
            if el is None:
                continue
            sys_clr = el.find("a:sysClr", ns)
            srgb_clr = el.find("a:srgbClr", ns)
            if srgb_clr is not None:
                colors[i] = srgb_clr.get("val", colors[i])
            elif sys_clr is not None:
                colors[i] = sys_clr.get("lastClr", colors[i])
        return colors
    except Exception:
        return list(DEFAULT_THEME_COLORS)


def _resolve_color(color_obj, theme_colors: List[str]) -> Optional[str]:
    """将 openpyxl Color 对象解析为 #RRGGBB 字符串"""
    if color_obj is None:
        return None
    try:
        color_type = getattr(color_obj, "type", None)
        tint = getattr(color_obj, "tint", 0.0) or 0.0

        if color_type == "rgb" or (color_type is None and getattr(color_obj, "rgb", None)):
            rgb = color_obj.rgb
            if rgb and isinstance(rgb, str) and len(rgb) >= 6:
                hex_rgb = rgb[-6:]
                if hex_rgb == "000000" and len(rgb) == 8 and rgb[:2] == "00":
                    return None  # 全透明
                hex_rgb = _apply_tint(hex_rgb, tint)
                return f"#{hex_rgb}"

        if color_type == "theme":
            theme_idx = getattr(color_obj, "theme", None)
            if theme_idx is not None and 0 <= theme_idx < len(theme_colors):
                hex_rgb = theme_colors[theme_idx]
                hex_rgb = _apply_tint(hex_rgb, tint)
                return f"#{hex_rgb}"

        if color_type == "indexed":
            idx = getattr(color_obj, "indexed", None)
            if idx is not None and 0 <= idx < len(INDEXED_COLORS):
                hex_rgb = INDEXED_COLORS[idx]
                hex_rgb = _apply_tint(hex_rgb, tint)
                return f"#{hex_rgb}"

    except Exception:
        pass
    return None


# ============================================================
# 边框样式映射
# ============================================================
BORDER_STYLE_MAP = {
    None: "none",
    "none": "none",
    "thin": "1px solid",
    "medium": "2px solid",
    "thick": "3px solid",
    "double": "3px double",
    "dotted": "1px dotted",
    "dashed": "1px dashed",
    "hair": "1px solid",
    "dashDot": "1px dashed",
    "dashDotDot": "1px dashed",
    "mediumDashed": "2px dashed",
    "mediumDashDot": "2px dashed",
    "mediumDashDotDot": "2px dashed",
    "slantDashDot": "2px dashed",
}


def _border_css(side, theme_colors: List[str]) -> Optional[str]:
    """将 openpyxl Border Side 转成 CSS border 值"""
    if side is None or side.style is None or side.style == "none":
        return None
    css_style = BORDER_STYLE_MAP.get(side.style, "1px solid")
    color = _resolve_color(side.color, theme_colors) if side.color else None
    if not color:
        color = "#000"
    return f"{css_style} {color}"


# ============================================================
# 单元格样式提取
# ============================================================

def _is_light_color(hex_color: str) -> bool:
    """判断颜色是否过浅（亮度 > 0.85），在白底上几乎不可见"""
    hex_rgb = hex_color.lstrip("#")
    if len(hex_rgb) != 6:
        return False
    r = int(hex_rgb[0:2], 16) / 255.0
    g = int(hex_rgb[2:4], 16) / 255.0
    b = int(hex_rgb[4:6], 16) / 255.0
    # 感知亮度公式
    luminance = 0.299 * r + 0.587 * g + 0.114 * b
    return luminance > 0.85


def _cell_style_css(cell, theme_colors: List[str]) -> str:
    """从 openpyxl cell 提取格式并返回 CSS style 字符串"""
    parts = []
    fg_color = None
    bg_color = None

    # --- 字体 ---
    font = cell.font
    if font:
        if font.bold:
            parts.append("font-weight:bold")
        if font.italic:
            parts.append("font-style:italic")

        decorations = []
        if font.underline and font.underline != "none":
            decorations.append("underline")
        if font.strikethrough:
            decorations.append("line-through")
        if decorations:
            parts.append(f"text-decoration:{' '.join(decorations)}")

        if font.size:
            parts.append(f"font-size:{font.size}pt")

        if font.name:
            parts.append(f"font-family:'{font.name}',sans-serif")

        fg_color = _resolve_color(font.color, theme_colors)

    # --- 背景色 ---
    fill = cell.fill
    if fill:
        if fill.patternType and fill.patternType != "none":
            bg_color = _resolve_color(fill.fgColor, theme_colors)
        if bg_color and bg_color.upper() != "#FFFFFF":
            parts.append(f"background:{bg_color}")
        else:
            bg_color = None  # 白色背景等同于无背景

    # --- 字体颜色（兜底：浅色文字 + 无深色背景 → 强制深色） ---
    if fg_color and fg_color.upper() != "#000000":
        if bg_color or not _is_light_color(fg_color):
            # 有深色背景，或者字体本身不是浅色，正常输出
            parts.append(f"color:{fg_color}")
        # else: 浅色字 + 白底 → 不输出 color，让它保持默认黑色

    # --- 对齐 ---
    alignment = cell.alignment
    if alignment:
        h_align = alignment.horizontal
        if h_align and h_align != "general":
            align_map = {"left": "left", "center": "center", "right": "right",
                         "justify": "justify", "distributed": "justify",
                         "centerContinuous": "center", "fill": "left"}
            css_align = align_map.get(h_align)
            if css_align:
                parts.append(f"text-align:{css_align}")

        v_align = alignment.vertical
        if v_align:
            v_map = {"top": "top", "center": "middle", "bottom": "bottom",
                     "justify": "top", "distributed": "top"}
            css_v = v_map.get(v_align)
            if css_v:
                parts.append(f"vertical-align:{css_v}")

        if alignment.wrap_text:
            parts.append("white-space:pre-wrap")

    # --- 边框 ---
    border = cell.border
    if border:
        bt = _border_css(border.top, theme_colors)
        bb = _border_css(border.bottom, theme_colors)
        bl = _border_css(border.left, theme_colors)
        br = _border_css(border.right, theme_colors)
        if bt:
            parts.append(f"border-top:{bt}")
        if bb:
            parts.append(f"border-bottom:{bb}")
        if bl:
            parts.append(f"border-left:{bl}")
        if br:
            parts.append(f"border-right:{br}")

    return ";".join(parts)


# ============================================================
# 工具函数
# ============================================================

def _safe_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    return html.escape(text).replace("\n", "<br>")


def _is_nonempty(value) -> bool:
    if value is None:
        return False
    return str(value).strip() != ""


def _collect_images(ws) -> Dict[Tuple[int, int], List[str]]:
    images: Dict[Tuple[int, int], List[str]] = {}
    for image in getattr(ws, "_images", []):
        try:
            anchor = image.anchor
            row = anchor._from.row + 1
            col = anchor._from.col + 1
            data = image._data()
            fmt = (getattr(image, "format", None) or "png").lower()
            b64 = base64.b64encode(data).decode("ascii")
            tag = f'<img src="data:image/{fmt};base64,{b64}" class="excel-image" />'
            images.setdefault((row, col), []).append(tag)
        except Exception:
            continue
    return images


def _build_merge_maps(ws, keep_rows: set, keep_cols: set):
    anchor_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
    covered_map: Dict[Tuple[int, int], Tuple[int, int]] = {}

    for merged in ws.merged_cells.ranges:
        min_row, max_row = merged.min_row, merged.max_row
        min_col, max_col = merged.min_col, merged.max_col

        rows_in = [r for r in range(min_row, max_row + 1) if r in keep_rows]
        cols_in = [c for c in range(min_col, max_col + 1) if c in keep_cols]
        if not rows_in or not cols_in:
            continue

        anchor_row = rows_in[0]
        anchor_col = cols_in[0]
        rowspan = len(rows_in)
        colspan = len(cols_in)
        anchor_map[(anchor_row, anchor_col)] = (rowspan, colspan)

        for r in rows_in:
            for c in cols_in:
                if r == anchor_row and c == anchor_col:
                    continue
                covered_map[(r, c)] = (anchor_row, anchor_col)

    return anchor_map, covered_map


def _excel_col_width_to_px(width: Optional[float]) -> Optional[int]:
    if width is None:
        return None
    try:
        w = float(width)
    except (TypeError, ValueError):
        return None
    if w <= 0:
        return None
    return max(8, int(round(w * 7 + 5)))


def _excel_row_height_to_px(height_pt: Optional[float]) -> Optional[int]:
    if height_pt is None:
        return None
    try:
        h = float(height_pt)
    except (TypeError, ValueError):
        return None
    if h <= 0:
        return None
    return max(8, int(round(h * 96 / 72)))


# ============================================================
# Sheet -> HTML
# ============================================================

def _sheet_to_html(ws, sheet_name: str, theme_colors: List[str]) -> str:
    from openpyxl.utils import get_column_letter

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return ""

    image_map = _collect_images(ws)

    keep_rows = set()
    keep_cols = set()

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            has_img = (r, c) in image_map
            if _is_nonempty(val) or has_img:
                keep_rows.add(r)
                keep_cols.add(c)

    for merged in ws.merged_cells.ranges:
        min_row, max_row_m = merged.min_row, merged.max_row
        min_col, max_col_m = merged.min_col, merged.max_col
        anchor_val = ws.cell(min_row, min_col).value
        has_img = (min_row, min_col) in image_map
        if _is_nonempty(anchor_val) or has_img:
            keep_rows.update(range(min_row, max_row_m + 1))
            keep_cols.update(range(min_col, max_col_m + 1))

    if not keep_rows or not keep_cols:
        return ""

    rows = list(range(min(keep_rows), max(keep_rows) + 1))
    cols = list(range(min(keep_cols), max(keep_cols) + 1))

    anchor_map, covered_map = _build_merge_maps(ws, set(rows), set(cols))

    lines = []
    lines.append('<section class="sheet">')
    lines.append(f"<h2>{html.escape(sheet_name)}</h2>")
    lines.append('<div class="xlsx-wrap">')
    lines.append('<table class="xlsx-table">')
    lines.append("<colgroup>")
    for c in cols:
        col_letter = get_column_letter(c)
        col_dim = ws.column_dimensions.get(col_letter)
        width_px = _excel_col_width_to_px(col_dim.width if col_dim else None)
        if width_px:
            lines.append(f'<col style="width:{width_px}px;">')
        else:
            lines.append("<col>")
    lines.append("</colgroup>")

    for r in rows:
        row_dim = ws.row_dimensions.get(r)
        row_height_px = _excel_row_height_to_px(row_dim.height if row_dim else None)
        if row_height_px:
            lines.append(f'<tr style="height:{row_height_px}px;">')
        else:
            lines.append("<tr>")
        for c in cols:
            if (r, c) in covered_map:
                continue

            cell = ws.cell(row=r, column=c)
            val = cell.value
            text = _safe_text(val)
            images = image_map.get((r, c), [])
            content = text
            if images:
                image_html = "".join(images)
                content = f"{content}<div class='excel-images'>{image_html}</div>" if content else image_html

            # 构建 HTML 属性
            html_attrs = []
            span = anchor_map.get((r, c))
            if span:
                rowspan, colspan = span
                if rowspan > 1:
                    html_attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    html_attrs.append(f'colspan="{colspan}"')

            # 单元格样式
            cell_css = _cell_style_css(cell, theme_colors)
            if cell_css:
                html_attrs.append(f'style="{cell_css}"')

            attr_text = f" {' '.join(html_attrs)}" if html_attrs else ""
            lines.append(f"<td{attr_text}>{content}</td>")
        lines.append("</tr>")

    lines.append("</table>")
    lines.append("</div>")
    lines.append("</section>")
    return "\n".join(lines)


# ============================================================
# 主转换入口
# ============================================================

def convert_xlsx_to_html_document(xlsx_path: Path, title: Optional[str] = None, page_style: str = "") -> str:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("缺少依赖 openpyxl，请先安装：pip install openpyxl") from exc

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    page_title = title or xlsx_path.stem
    theme_colors = _extract_theme_colors(wb)

    sections = []
    nav_links = []
    for idx, name in enumerate(wb.sheetnames, start=1):
        ws = wb[name]
        section_html = _sheet_to_html(ws, name, theme_colors)
        if not section_html:
            continue
        anchor = f"sheet-{idx}"
        nav_links.append(f'<a href="#{anchor}">{html.escape(name)}</a>')
        sections.append(f'<a id="{anchor}"></a>')
        sections.append(section_html)

    if not sections:
        sections.append('<p><em>该 Excel 未找到可显示内容。</em></p>')

    table_style = """
.sheet { margin: 16px 0 28px; }
.sheet h2 { margin: 8px 0; color: #2c3e50; font-size: 1.15em; }
.sheet-nav { margin: 8px 0 14px; display: flex; flex-wrap: wrap; gap: 8px; }
.sheet-nav a {
    display: inline-block;
    padding: 2px 8px;
    border: 1px solid #cfd8dc;
    border-radius: 6px;
    color: #1f4e79;
    text-decoration: none;
    background: #f7fbff;
}
.xlsx-wrap {
    overflow: auto;
    background: #fff;
    width: fit-content;
    max-width: 100%;
}
.xlsx-table { border-collapse: collapse; width: auto; table-layout: auto; }
.xlsx-table td {
    padding: 2px 4px;
    vertical-align: top;
    white-space: pre-wrap;
    word-break: break-word;
    line-height: 1.45;
    border: 1px solid #d0d0d0;
}
.excel-images { margin-top: 6px; }
.excel-image { max-width: 260px; max-height: 260px; height: auto; border: 1px solid #ddd; background: #fff; }
""".strip()

    style_block = page_style + "\n" + table_style if page_style else table_style
    nav_html = "" if not nav_links else f'<div class="sheet-nav">{"".join(nav_links)}</div>'

    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>{html.escape(page_title)}</title>
<style>{style_block}</style>
</head>
<body>
<h1>{html.escape(page_title)}</h1>
{nav_html}
{chr(10).join(sections)}
</body>
</html>"""


def main():
    parser = argparse.ArgumentParser(description="Excel(.xlsx) 转 HTML（适配 CHM 阅读）")
    parser.add_argument("input", help="输入 .xlsx 文件")
    parser.add_argument("-o", "--output", help="输出 .html 路径（默认同名）")
    parser.add_argument("--title", help="页面标题")
    parser.add_argument("--gbk", action="store_true", help="输出 GBK 编码（默认 UTF-8）")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else input_path.with_suffix(".html")

    html_content = convert_xlsx_to_html_document(input_path, title=args.title)
    if args.gbk:
        html_content = html_content.replace('charset="utf-8"', 'charset="gbk"')
        output_path.write_bytes(html_content.encode("gbk", errors="xmlcharrefreplace"))
    else:
        output_path.write_text(html_content, encoding="utf-8")

    print(f"输出: {output_path}")


if __name__ == "__main__":
    main()
